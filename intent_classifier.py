# ============================================
# SEO Automation — Episode 2
# Intent Classification — 51k Keywords
# Built by Santhosh Natarajan
# ============================================
# What this script does:
# - Reads your GSC data (51k keywords)
# - Classifies every keyword by intent:
#   💰 Transactional — ready to enroll/decide
#   🧲 Commercial — comparing & researching career
#   📚 Informational — learning something
#   🔗 Navigational — looking for specific page
# - Splits into separate sheets by intent
# - Color codes everything
# - Saves to Excel automatically
# ============================================

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────
# CONFIG — Change if needed
# ─────────────────────────────────────────
INPUT_FILE  = r"C:\seo-automation\gsc_data.xlsx"
OUTPUT_FILE = r"C:\seo-automation\intent_classified.xlsx"

# ─────────────────────────────────────────
# INTENT KEYWORDS — Tuned for Training Institute
# ─────────────────────────────────────────

TRANSACTIONAL = [
    # Location + Course = ready to enroll
    'in coimbatore', 'in chennai', 'in trichy', 'in madurai',
    'in tamil nadu', 'near me',

    # Money / decision signals
    'fees', 'fee', 'cost', 'price', 'duration',
    'syllabus', 'course fees', 'training fees',

    # Action signals
    'enroll', 'admission', 'register', 'join',
    'apply for', 'how to apply',

    # Outcome signals (job-focused decisions)
    'with placement', 'placement guarantee',
    'placement assistance', 'job guarantee',

    # Course-specific decision
    'best institute', 'best training center',
    'best coaching center', 'top institute',
]

COMMERCIAL = [
    # Comparison
    'best', 'top', 'top 10', 'top 5', 'list of',
    'vs', 'versus', 'compare', 'comparison',
    'difference between', 'which is better',
    'pros and cons', 'review', 'reviews',

    # Career research (pre-enrollment)
    'salary', 'jobs', 'job opportunities',
    'career in', 'career options', 'scope of',
    'future of', 'is it worth', 'demand for',
    'skills required', 'eligibility',

    # After-degree research
    'after bca', 'after mca', 'after bsc', 'after b.sc',
    'after engineering', 'after 12th', 'after degree',
    'after b.com', 'after mba', 'after bba',

    # Interview prep (job seekers)
    'interview questions', 'interview question',
]

INFORMATIONAL = [
    # Learning signals
    'what is', 'what are', 'how does', 'how do',
    'why is', 'why does', 'explain', 'meaning',
    'definition', 'introduction',

    # Tutorial signals
    'tutorial', 'guide', 'learn', 'basics',
    'example', 'examples', 'types of',

    # Programming queries (huge for Systech)
    'program', 'code', 'algorithm', 'in python',
    'in java', 'using python', 'using java',
    'series', 'write a', 'how to write',

    # How-to queries (technical)
    'how to create', 'how to build', 'how to make',
    'how to use', 'how to install', 'how to setup',
    'how to configure', 'how to enable', 'how to connect',
    'how to login', 'how to change', 'how to fix',
    'how to reset', 'how to find',

    # Technical reference (router/password queries)
    'login', 'username', 'password', 'admin login',
    'router login', 'password change', 'gateway',
    'ip address', 'public ip', 'my ip',
]

# ─────────────────────────────────────────
# CLASSIFY INTENT
# ─────────────────────────────────────────
def classify_intent(query):
    q = str(query).lower()

    # Transactional first (highest priority)
    for kw in TRANSACTIONAL:
        if kw in q:
            return 'Transactional'

    # Then Commercial
    for kw in COMMERCIAL:
        if kw in q:
            return 'Commercial'

    # Then Informational
    for kw in INFORMATIONAL:
        if kw in q:
            return 'Informational'

    # Default
    if len(q.split()) <= 2:
        return 'Navigational'

    return 'Informational'

# ─────────────────────────────────────────
# COLOR MAP
# ─────────────────────────────────────────
COLORS = {
    'Transactional': '10B981',  # Green
    'Commercial':    'F59E0B',  # Amber
    'Informational': '3B82F6',  # Blue
    'Navigational':  '8B5CF6',  # Purple
}

ICONS = {
    'Transactional': '💰',
    'Commercial':    '🧲',
    'Informational': '📚',
    'Navigational':  '🔗',
}

# ─────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────
def style_header(ws, color, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(name="Arial", bold=True,
                        color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal='center',
                                  vertical='center')
    ws.row_dimensions[1].height = 30

def style_rows(ws, intent, num_cols):
    color = COLORS.get(intent, '6B7280')
    light_colors = {
        'Transactional': 'D1FAE5',
        'Commercial':    'FEF3C7',
        'Informational': 'DBEAFE',
        'Navigational':  'EDE9FE',
    }
    light = light_colors.get(intent, 'F9FAFB')

    for row in range(2, ws.max_row + 1):
        bg = light if row % 2 == 0 else 'FFFFFF'
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical='center')

        # Color the Intent column
        intent_cell = ws.cell(row=row, column=num_cols)
        intent_cell.fill = PatternFill("solid", fgColor=color)
        intent_cell.font = Font(name="Arial", bold=True,
                               color="FFFFFF", size=10)
        intent_cell.alignment = Alignment(horizontal='center',
                                         vertical='center')

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 70)

# ─────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 50)
    print("  Intent Classifier — Santhosh Natarajan")
    print("  SEO Automation Series — Episode 2")
    print("=" * 50)

    # Load GSC data
    print("\n📂 Loading GSC data...")
    df = pd.read_excel(INPUT_FILE)
    print(f"✅ Loaded {len(df):,} keywords")

    # Classify intent
    print("\n🧠 Classifying intent for all keywords...")
    df['Intent'] = df['Query'].apply(classify_intent)
    print("✅ Classification complete")

    # Summary
    print("\n📊 Intent Summary:")
    summary = df['Intent'].value_counts()
    for intent, count in summary.items():
        icon = ICONS[intent]
        pct = round(count / len(df) * 100, 1)
        print(f"  {icon} {intent}: {count:,} keywords ({pct}%)")

    # Write to Excel
    print("\n💾 Writing to Excel...")

    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:

        # Sheet 1 — Full data
        df_export = df[['Query','Clicks','Impressions',
                        'CTR','Position','Intent']].copy()
        df_export.to_excel(writer, index=False,
                          sheet_name='📊 All Keywords')

        # Sheet 2-5 — One per intent
        for intent in ['Transactional','Commercial',
                       'Informational','Navigational']:
            icon = ICONS[intent]
            sheet_name = f"{icon} {intent}"
            filtered = df[df['Intent'] == intent][
                ['Query','Clicks','Impressions',
                 'CTR','Position','Intent']
            ].copy()
            filtered.to_excel(writer, index=False,
                            sheet_name=sheet_name)

        # Sheet 6 — Summary
        summary_data = []
        for intent in ['Transactional','Commercial',
                       'Informational','Navigational']:
            subset = df[df['Intent'] == intent]
            summary_data.append({
                'Intent':            f"{ICONS[intent]} {intent}",
                'Total Keywords':    len(subset),
                'Total Impressions': int(subset['Impressions'].sum()),
                'Total Clicks':      int(subset['Clicks'].sum()),
                'Avg Position':      round(
                                         subset['Position'].mean(), 1
                                     ) if len(subset) else 0,
                'What to do':        {
                    'Transactional': 'Optimize landing pages + CTAs',
                    'Commercial':    'Write comparison & review posts',
                    'Informational': 'Create tutorials & how-to guides',
                    'Navigational':  'Ensure correct pages exist',
                }[intent]
            })
        pd.DataFrame(summary_data).to_excel(
            writer, index=False, sheet_name='📋 Summary'
        )

    # Apply styles
    print("🎨 Applying colors and styles...")
    wb = load_workbook(OUTPUT_FILE)

    intent_colors = {
        '📊 All Keywords':    '1E1E2E',
        '💰 Transactional':   COLORS['Transactional'],
        '🧲 Commercial':      COLORS['Commercial'],
        '📚 Informational':   COLORS['Informational'],
        '🔗 Navigational':    COLORS['Navigational'],
        '📋 Summary':         '1E1E2E',
    }

    for sheet_name, color in intent_colors.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            num_cols = ws.max_column
            style_header(ws, color, num_cols)
            intent_key = sheet_name.split(' ', 1)[-1] if ' ' in sheet_name else None
            if intent_key in COLORS:
                style_rows(ws, intent_key, num_cols)
            auto_width(ws)
            ws.freeze_panes = 'A2'

    wb.save(OUTPUT_FILE)

    print(f"\n{'=' * 50}")
    print(f"✅ DONE — {len(df):,} keywords classified")
    print(f"📁 File: {OUTPUT_FILE}")
    print(f"{'=' * 50}")
    print("\n🚀 Episode 2 complete. Open the Excel file now.")